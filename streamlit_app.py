"""
CA Final Dashboard — Streamlit Version
Install: pip install streamlit plotly pandas openpyxl
Run:     streamlit run streamlit_app.py
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import warnings
warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CA Final Dashboard",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

EXCEL_FILE = "CA_Final_Tracker.xlsx"
EXAM_DATE  = date(2027, 1, 1)
SUBJECTS   = ["FR", "AFM", "AA", "DT", "IDT"]
SUBJ_FULL  = {
    "FR" : "Financial Reporting",
    "AFM": "Adv. FM & Economics",
    "AA" : "Advanced Auditing",
    "DT" : "Direct Tax & Int'l Tax",
    "IDT": "Indirect Tax"
}
TARGET_HRS = {"FR": 200, "AFM": 160, "AA": 150, "DT": 200, "IDT": 180}
COLORS     = {"FR":"#7C3AED","AFM":"#10B981","AA":"#F59E0B","DT":"#EF4444","IDT":"#3B82F6"}

TOPICS = {
"FR":["Ind AS 1 – Presentation of FS","Ind AS 2 – Inventories",
      "Ind AS 7 – Cash Flow Statements","Ind AS 8 – Accounting Policies",
      "Ind AS 10 – Events after Reporting Period","Ind AS 12 – Deferred Tax",
      "Ind AS 16 – Property Plant & Equipment","Ind AS 19 – Employee Benefits",
      "Ind AS 20 – Government Grants","Ind AS 21 – Foreign Currency",
      "Ind AS 23 – Borrowing Costs","Ind AS 24 – Related Party Disclosures",
      "Ind AS 27 – Separate Financial Statements","Ind AS 28 – Associates & JV",
      "Ind AS 32 – Financial Instruments: Presentation",
      "Ind AS 33 – Earnings per Share","Ind AS 36 – Impairment of Assets",
      "Ind AS 37 – Provisions & Contingencies","Ind AS 38 – Intangible Assets",
      "Ind AS 40 – Investment Property","Ind AS 101 – First-time Adoption",
      "Ind AS 102 – Share-based Payments","Ind AS 103 – Business Combinations",
      "Ind AS 105 – Assets Held for Sale","Ind AS 108 – Operating Segments",
      "Ind AS 109 – Financial Instruments","Ind AS 110 – Consolidated FS",
      "Ind AS 111 – Joint Arrangements","Ind AS 113 – Fair Value Measurement",
      "Ind AS 115 – Revenue from Contracts","Ind AS 116 – Leases",
      "Analysis & Interpretation of FS"],
"AFM":["Financial Policy & Corporate Strategy","Risk Management – Overview",
       "Capital Budgeting under Risk & Uncertainty","Dividend Policy",
       "Indian Capital Market & SEBI","Security Analysis – Fundamental & Technical",
       "Portfolio Management & CAPM","Mutual Funds",
       "Derivatives – Futures & Forwards","Derivatives – Options",
       "Derivatives – Swaps & Interest Rate","Foreign Exchange Risk Management",
       "International Financial Management","Mergers Acquisitions & Restructuring",
       "Startup Finance & Venture Capital","Leasing & Hire Purchase",
       "Bond Valuation & Interest Rate Risk","Economic Value Added (EVA)",
       "Financial Modelling & Simulation"],
"AA":["Nature Objective & Scope of Audit","Ethics & Independence (SA 200-299)",
      "Audit Planning Materiality & Risk","Internal Control & Internal Audit",
      "Audit Evidence – SA 500 series","Sampling & CAAT",
      "Verification of Assets & Liabilities","Company Audit – Specific Areas",
      "Audit Report & Modified Opinions","Special Audits – Banks Insurance NBFCs",
      "Cost Audit","Forensic Accounting & Fraud Investigation",
      "Peer Review & Quality Control (SQC 1)","Audit under IT Environment",
      "Concurrent & Revenue Audit","Due Diligence & Investigations"],
"DT":["Basic Concepts & Residential Status","Incomes Exempt from Tax",
      "Income from Salaries","Income from House Property",
      "Profits & Gains – Business/Profession","Capital Gains",
      "Income from Other Sources","Clubbing Set-off & Carry Forward",
      "Deductions under Chapter VIA","Assessment – Individuals HUF Firms",
      "Assessment – Companies & Other Entities","MAT & AMT",
      "TDS & TCS Provisions","Advance Tax & Interest",
      "Return Filing & Assessment Procedure","Appeals & Revision",
      "International Taxation – Transfer Pricing","DTAA & OECD/UN Model",
      "GAAR POEM & BEPS"],
"IDT":["GST – Constitutional Background","GST – Levy & Exemptions",
       "GST – Time Place & Value of Supply","GST – Input Tax Credit",
       "GST – Registration","GST – Tax Invoice Credit & Debit Notes",
       "GST – Returns","GST – Payment & Refund",
       "GST – Import & Export (Zero-rated)","GST – Job Work & E-Commerce",
       "GST – Assessment & Audit","GST – Demand Adjudication & Recovery",
       "GST – Appeals & Revision","GST – Offences & Penalties",
       "GST – Miscellaneous Provisions","Customs – Levy & Exemptions",
       "Customs – Import/Export Procedure","Customs – Valuation & Baggage Rules",
       "Customs – Refund Drawback & Special Provisions","FTP – Overview"]
}

# ── CUSTOM CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #1E1E2E; }
    .stApp { background-color: #1E1E2E; }
    [data-testid="stSidebar"] { background-color: #2D2D3F; }
    .metric-card {
        background: #2D2D3F;
        border-radius: 12px;
        padding: 16px 20px;
        margin: 4px;
        border-left: 4px solid;
    }
    .success-msg {
        background: #064E3B;
        color: #6EE7B7;
        padding: 10px 16px;
        border-radius: 8px;
        margin: 8px 0;
    }
    div[data-testid="stMetric"] {
        background: #2D2D3F;
        border-radius: 12px;
        padding: 12px 16px;
    }
    .stSelectbox > div, .stNumberInput > div {
        background: #2D2D3F !important;
    }
    h1,h2,h3,p,label { color: #E2E8F0 !important; }
    .stTabs [data-baseweb="tab"] { color: #94A3B8; }
    .stTabs [aria-selected="true"] { color: #7C3AED !important; }
</style>
""", unsafe_allow_html=True)

# ── DATA LOADER ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=30)
def load_data():
    try:
        log = pd.read_excel(EXCEL_FILE, sheet_name="Daily_Log",        header=1)
        rev = pd.read_excel(EXCEL_FILE, sheet_name="Revision_Tracker", header=1)
        tst = pd.read_excel(EXCEL_FILE, sheet_name="Test_Scores",      header=1)
        top = pd.read_excel(EXCEL_FILE, sheet_name="Topics_Master",    header=1)
    except Exception as e:
        return None, None, None, None, str(e)

    # Clean Daily Log
    log.columns = [str(c).strip() for c in log.columns]
    log = log.dropna(subset=["Date","Subject","Hours"])
    log["Date"]    = pd.to_datetime(log["Date"], errors="coerce")
    log["Hours"]   = pd.to_numeric(log["Hours"], errors="coerce").fillna(0)
    log["Subject"] = log["Subject"].astype(str).str.strip().str.upper()
    log = log[log["Subject"].isin(SUBJECTS)]

    # Clean Test Scores
    tst.columns  = [str(c).strip() for c in tst.columns]
    tst = tst.dropna(subset=["Marks","Max Marks"])
    tst["Marks"]     = pd.to_numeric(tst["Marks"],      errors="coerce")
    tst["Max Marks"] = pd.to_numeric(tst["Max Marks"],  errors="coerce")
    tst["Score %"]   = (tst["Marks"]/tst["Max Marks"]*100).round(1)
    tst["Date"]      = pd.to_datetime(tst["Date"],      errors="coerce")
    tst["Subject"]   = tst["Subject"].astype(str).str.strip().str.upper()

    # Clean Revision
    rev.columns  = [str(c).strip() for c in rev.columns]
    rev["Subject"] = rev["Subject"].astype(str).str.strip().str.upper()

    return log, rev, tst, top, None

# ── EXCEL WRITER ──────────────────────────────────────────────────────────────
def append_to_excel(sheet_name, row_data):
    """Append a row to the specified sheet"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        # Find next empty row (skip header rows)
        next_row = ws.max_row + 1
        for col, val in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=val)
        wb.save(EXCEL_FILE)
        return True, "✅ Saved successfully!"
    except Exception as e:
        return False, f"❌ Error: {e}"

def update_revision(subject, topic, field, value):
    """Update a specific cell in revision tracker"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Revision_Tracker"]
        col_map = {
            "1st Read": 3, "R1 Date": 4,
            "R2 Date": 5, "R3 Date": 6,
            "Confidence(1-5)": 7, "Due Revision": 8, "Notes": 9
        }
        col = col_map.get(field)
        if not col:
            return False, "Invalid field"
        for row in range(3, ws.max_row + 1):
            s = str(ws.cell(row=row, column=1).value).strip().upper()
            t = str(ws.cell(row=row, column=2).value).strip()
            if s == subject.upper() and t == topic:
                ws.cell(row=row, column=col).value = value
                wb.save(EXCEL_FILE)
                return True, "✅ Revision updated!"
        return False, "❌ Topic not found"
    except Exception as e:
        return False, f"❌ Error: {e}"

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🎓 CA Final Tracker")
    st.markdown("---")
    page = st.radio("Navigate", [
        "📊 Dashboard",
        "📝 Log Today's Study",
        "🏆 Add Test Score",
        "🔄 Update Revision",
        "📋 View All Data"
    ], label_visibility="collapsed")
    st.markdown("---")
    days_left = max((EXAM_DATE - date.today()).days, 0)
    st.markdown(f"### ⏳ {days_left} days to exam")
    progress = max(0, min(1, 1 - days_left/365))
    st.progress(progress)
    st.caption(f"Exam: January 2027")
    st.markdown("---")
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
log, rev, tst, top, err = load_data()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if page == "📊 Dashboard":
    st.title("📊 CA Final — Performance Dashboard")

    if err:
        st.error(f"Could not load Excel file: {err}")
        st.info("Make sure CA_Final_Tracker.xlsx exists. Run build_tracker.py first.")
        st.stop()

    # ── KPI Row
    total_hrs = log["Hours"].sum() if log is not None else 0
    total_tgt = sum(TARGET_HRS.values())
    avg_score = tst["Score %"].mean() if tst is not None and not tst.empty else 0
    sh = log.groupby("Subject")["Hours"].sum() if log is not None else pd.Series()
    need = max(total_tgt - total_hrs, 0)
    dpd  = round(need / days_left, 1) if days_left > 0 else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("⏳ Days Left",       f"{days_left}",       "to Jan 2027")
    c2.metric("📚 Total Hours",     f"{total_hrs:.0f}h",  f"{dpd}h/day needed")
    c3.metric("🎯 Avg Test Score",  f"{avg_score:.1f}%",  "Target: 60%+")
    c4.metric("📈 Sessions Logged", f"{len(log)}",        "total sessions")
    c5.metric("✅ Target Hours",    f"{total_tgt}h",      f"{total_hrs:.0f}h done")

    st.markdown("---")

    # ── Subject Progress Bars
    st.subheader("📚 Subject Progress")
    cols = st.columns(5)
    for i, s in enumerate(SUBJECTS):
        done = sh.get(s, 0)
        tgt  = TARGET_HRS[s]
        pct  = min(done/tgt*100, 100)
        with cols[i]:
            st.markdown(f"**{s}**")
            st.progress(int(pct))
            st.caption(f"{done:.0f}h / {tgt}h ({pct:.0f}%)")

    st.markdown("---")

    # ── Charts Row 1
    col1, col2 = st.columns([2, 1])

    with col1:
        # Daily hours stacked bar (last 30 days)
        end   = date.today()
        start = end - timedelta(days=29)
        daily = log[log["Date"].dt.date >= start].copy()
        if not daily.empty:
            daily_grp = daily.groupby([daily["Date"].dt.date, "Subject"])["Hours"].sum().reset_index()
            daily_grp.columns = ["Date","Subject","Hours"]
            fig = px.bar(daily_grp, x="Date", y="Hours", color="Subject",
                         color_discrete_map=COLORS, barmode="stack",
                         title="📆 Daily Study Hours — Last 30 Days")
            fig.add_hline(y=6, line_dash="dash", line_color="#F59E0B",
                          annotation_text="6h target")
            fig.update_layout(paper_bgcolor="#2D2D3F", plot_bgcolor="#2D2D3F",
                              font_color="#E2E8F0", legend=dict(bgcolor="#2D2D3F"))
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        # Subject hours vs target
        sh_df = pd.DataFrame({
            "Subject": SUBJECTS,
            "Done":    [sh.get(s,0) for s in SUBJECTS],
            "Target":  [TARGET_HRS[s] for s in SUBJECTS]
        })
        fig2 = go.Figure()
        for _, row in sh_df.iterrows():
            pct = min(row["Done"]/row["Target"]*100,100)
            fig2.add_trace(go.Bar(
                x=[row["Done"]], y=[SUBJ_FULL[row["Subject"]]],
                orientation="h", name=row["Subject"],
                marker_color=COLORS[row["Subject"]],
                text=f"{row['Done']:.0f}h/{row['Target']}h",
                textposition="inside", showlegend=False
            ))
        fig2.update_layout(title="🎯 Hours vs Target",
                           paper_bgcolor="#2D2D3F", plot_bgcolor="#2D2D3F",
                           font_color="#E2E8F0",
                           xaxis=dict(range=[0, max(TARGET_HRS.values())+20]))
        st.plotly_chart(fig2, use_container_width=True)

    # ── Charts Row 2
    col3, col4 = st.columns([2,1])

    with col3:
        if tst is not None and not tst.empty:
            tst_s = tst.sort_values("Date")
            fig3  = go.Figure()
            for s in SUBJECTS:
                df = tst_s[tst_s["Subject"]==s]
                if df.empty: continue
                fig3.add_trace(go.Scatter(
                    x=df["Date"], y=df["Score %"],
                    name=SUBJ_FULL.get(s,s), mode="lines+markers",
                    line=dict(color=COLORS[s], width=2), marker=dict(size=8)
                ))
            fig3.add_hline(y=50, line_dash="dash", line_color="#EF4444",
                           annotation_text="Pass Line 50%")
            fig3.add_hline(y=60, line_dash="dot",  line_color="#10B981",
                           annotation_text="Target 60%")
            fig3.update_layout(title="📈 Test Score Trends",
                               paper_bgcolor="#2D2D3F", plot_bgcolor="#2D2D3F",
                               font_color="#E2E8F0", yaxis=dict(range=[0,105]),
                               legend=dict(bgcolor="#2D2D3F"))
            st.plotly_chart(fig3, use_container_width=True)

    with col4:
        if tst is not None and not tst.empty:
            by_s = tst.groupby("Subject")["Score %"].mean().reindex(SUBJECTS).fillna(0)
            clrs = ["#EF4444" if v<50 else ("#F59E0B" if v<60 else "#10B981")
                    for v in by_s.values]
            fig4 = go.Figure(go.Bar(
                x=by_s.index, y=by_s.values,
                marker_color=clrs,
                text=[f"{v:.1f}%" for v in by_s.values],
                textposition="outside"
            ))
            fig4.add_hline(y=50, line_dash="dash", line_color="#EF4444")
            fig4.add_hline(y=60, line_dash="dot",  line_color="#10B981")
            fig4.update_layout(title="🎯 Avg Score by Subject",
                               paper_bgcolor="#2D2D3F", plot_bgcolor="#2D2D3F",
                               font_color="#E2E8F0", yaxis=dict(range=[0,105]))
            st.plotly_chart(fig4, use_container_width=True)

    # ── Revision Status
    st.subheader("🔄 Revision Status")
    if rev is not None and not rev.empty:
        fig5 = make_subplots(rows=1, cols=5,
                             specs=[[{"type":"pie"}]*5],
                             subplot_titles=list(SUBJ_FULL.values()))
        for i, s in enumerate(SUBJECTS, 1):
            df    = rev[rev["Subject"]==s]
            total = len(df)
            if total == 0: continue
            r3 = df["R3 Date"].notna().sum() if "R3 Date" in df.columns else 0
            r2 = max(df["R2 Date"].notna().sum() - r3, 0) if "R2 Date" in df.columns else 0
            r1 = max(df["R1 Date"].notna().sum() - r3 - r2, 0) if "R1 Date" in df.columns else 0
            rd = max(df["1st Read"].isin(["Done","✓","done"]).sum() - r3-r2-r1, 0) if "1st Read" in df.columns else 0
            ns = max(total - r3 - r2 - r1 - rd, 0)
            fig5.add_trace(go.Pie(
                values=[r3,r2,r1,rd,ns],
                labels=["R3 Done","R2 Done","R1 Done","1st Read","Not Started"],
                marker_colors=["#10B981","#3B82F6","#F59E0B","#7C3AED","#4B5563"],
                hole=0.5, showlegend=(i==1), textinfo="percent"
            ), row=1, col=i)
        fig5.update_layout(paper_bgcolor="#2D2D3F", font_color="#E2E8F0",
                           legend=dict(bgcolor="#2D2D3F"), height=300)
        st.plotly_chart(fig5, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — LOG TODAY'S STUDY
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📝 Log Today's Study":
    st.title("📝 Log Today's Study Session")
    st.markdown("Fill in what you studied today — takes less than 1 minute!")

    with st.form("study_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            study_date  = st.date_input("📅 Date", value=date.today())
            subject     = st.selectbox("📚 Subject", SUBJECTS,
                                       format_func=lambda x: f"{x} — {SUBJ_FULL[x]}")
            hours       = st.number_input("⏱️ Hours Studied", 0.5, 12.0, 2.0, 0.5)
        with c2:
            topic       = st.selectbox("📖 Topic", TOPICS.get(subject, []))
            pages       = st.number_input("📄 Pages / Questions Done", 0, 500, 20, 5)
            difficulty  = st.select_slider("💪 Difficulty", [1,2,3,4,5],
                                           value=3,
                                           format_func=lambda x:
                                           ["","⭐ Very Easy","⭐⭐ Easy",
                                            "⭐⭐⭐ Medium","⭐⭐⭐⭐ Hard",
                                            "⭐⭐⭐⭐⭐ Very Hard"][x])
        notes = st.text_area("📝 Notes (optional)", placeholder="Key points, doubts, etc.")

        submitted = st.form_submit_button("✅ Save Study Session", use_container_width=True)

        if submitted:
            # Find next session number
            session_num = len(log) + 1 if log is not None else 1
            row = [study_date.strftime("%d-%b-%Y"), subject, topic,
                   hours, pages, difficulty, notes, session_num]
            ok, msg = append_to_excel("Daily_Log", row)
            if ok:
                st.success(msg)
                st.balloons()
                st.cache_data.clear()
            else:
                st.error(msg)

    # Show recent logs
    if log is not None and not log.empty:
        st.markdown("---")
        st.subheader("📋 Recent Study Sessions")
        recent = log.sort_values("Date", ascending=False).head(10).copy()
        recent["Date"] = recent["Date"].dt.strftime("%d %b %Y")
        st.dataframe(recent[["Date","Subject","Topic Studied","Hours",
                              "Pages/Qs","Difficulty (1-5)"]].reset_index(drop=True),
                     use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 3 — ADD TEST SCORE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🏆 Add Test Score":
    st.title("🏆 Add Mock Test / Practice Score")

    with st.form("test_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            test_date   = st.date_input("📅 Test Date", value=date.today())
            subject     = st.selectbox("📚 Subject", SUBJECTS + ["All"],
                                       format_func=lambda x:
                                       f"{x} — {SUBJ_FULL.get(x,'Full Syllabus')}")
            test_name   = st.text_input("📝 Test Name / Source",
                                        placeholder="e.g. ICAI Mock 1, Module Test")
        with c2:
            marks       = st.number_input("✅ Marks Obtained", 0, 200, 55)
            max_marks   = st.number_input("📊 Maximum Marks",  0, 200, 100)
            pct         = round(marks/max_marks*100,1) if max_marks > 0 else 0
            color       = "🟢" if pct>=60 else ("🟡" if pct>=50 else "🔴")
            st.metric("Score %", f"{color} {pct}%",
                      "Pass ✅" if pct>=50 else "Below Pass ❌")

        c3, c4 = st.columns(2)
        with c3:
            weak  = st.text_area("❌ Weak Areas", placeholder="Topics to revise...")
        with c4:
            strong= st.text_area("✅ Strong Areas", placeholder="Topics you aced...")
        action = st.text_area("📌 Action Plan",  placeholder="What will you do to improve?")

        submitted = st.form_submit_button("✅ Save Test Score", use_container_width=True)
        if submitted:
            row = [test_date.strftime("%d-%b-%Y"), subject, test_name,
                   marks, max_marks, pct,
                   "Yes" if pct>=50 else "No", weak, strong, action]
            ok, msg = append_to_excel("Test_Scores", row)
            if ok:
                st.success(msg)
                st.balloons()
                st.cache_data.clear()
            else:
                st.error(msg)

    # Show recent scores
    if tst is not None and not tst.empty:
        st.markdown("---")
        st.subheader("📋 Recent Test Scores")
        recent = tst.sort_values("Date", ascending=False).head(10).copy()
        recent["Date"] = recent["Date"].dt.strftime("%d %b %Y")
        st.dataframe(recent[["Date","Subject","Test Name / Source",
                              "Marks","Max Marks","Score %"]].reset_index(drop=True),
                     use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 4 — UPDATE REVISION
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🔄 Update Revision":
    st.title("🔄 Update Revision Status")

    c1, c2 = st.columns(2)
    with c1:
        subject = st.selectbox("📚 Subject", SUBJECTS,
                               format_func=lambda x: f"{x} — {SUBJ_FULL[x]}")
    with c2:
        topic = st.selectbox("📖 Topic", TOPICS.get(subject, []))

    st.markdown("---")
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("### 📖 First Read")
        if st.button("✅ Mark as Done", key="r0", use_container_width=True):
            ok, msg = update_revision(subject, topic, "1st Read", "Done")
            st.success(msg) if ok else st.error(msg)
            st.cache_data.clear()

    with col2:
        st.markdown("### 🔄 Revision Dates")
        r1_date = st.date_input("R1 Date", key="r1d")
        if st.button("💾 Save R1", key="r1", use_container_width=True):
            ok, msg = update_revision(subject, topic, "R1 Date",
                                      r1_date.strftime("%d-%b-%Y"))
            st.success(msg) if ok else st.error(msg)
            st.cache_data.clear()

        r2_date = st.date_input("R2 Date", key="r2d")
        if st.button("💾 Save R2", key="r2", use_container_width=True):
            ok, msg = update_revision(subject, topic, "R2 Date",
                                      r2_date.strftime("%d-%b-%Y"))
            st.success(msg) if ok else st.error(msg)
            st.cache_data.clear()

        r3_date = st.date_input("R3 Date", key="r3d")
        if st.button("💾 Save R3", key="r3", use_container_width=True):
            ok, msg = update_revision(subject, topic, "R3 Date",
                                      r3_date.strftime("%d-%b-%Y"))
            st.success(msg) if ok else st.error(msg)
            st.cache_data.clear()

    with col3:
        st.markdown("### ⭐ Confidence")
        conf = st.select_slider("Rate yourself", [1,2,3,4,5],
                                format_func=lambda x:
                                ["","😰 1","😕 2","😐 3","😊 4","🔥 5"][x])
        if st.button("💾 Save Confidence", use_container_width=True):
            ok, msg = update_revision(subject, topic, "Confidence(1-5)", conf)
            st.success(msg) if ok else st.error(msg)
            st.cache_data.clear()

        due = st.selectbox("Due for Revision?", ["Yes","No","Soon"])
        if st.button("💾 Save Due Status", use_container_width=True):
            ok, msg = update_revision(subject, topic, "Due Revision", due)
            st.success(msg) if ok else st.error(msg)
            st.cache_data.clear()

    # Show current revision status for selected subject
    if rev is not None and not rev.empty:
        st.markdown("---")
        st.subheader(f"📋 {subject} — Revision Status")
        df = rev[rev["Subject"]==subject].copy()
        if not df.empty:
            cols_show = ["Topic","1st Read","R1 Date","R2 Date",
                         "R3 Date","Confidence(1-5)","Due Revision"]
            cols_show = [c for c in cols_show if c in df.columns]
            st.dataframe(df[cols_show].reset_index(drop=True),
                         use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 5 — VIEW ALL DATA
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📋 View All Data":
    st.title("📋 View All Data")

    tab1, tab2, tab3 = st.tabs(["📚 Daily Log","🏆 Test Scores","🔄 Revision"])

    with tab1:
        if log is not None and not log.empty:
            subj_filter = st.multiselect("Filter by Subject",SUBJECTS,default=SUBJECTS)
            filtered = log[log["Subject"].isin(subj_filter)].copy()
            filtered["Date"] = filtered["Date"].dt.strftime("%d %b %Y")
            st.dataframe(filtered.reset_index(drop=True), use_container_width=True)
            st.caption(f"Total: {len(filtered)} sessions | {filtered['Hours'].sum():.1f} hours")

    with tab2:
        if tst is not None and not tst.empty:
            st.dataframe(tst.reset_index(drop=True), use_container_width=True)

    with tab3:
        if rev is not None and not rev.empty:
            subj_filter2 = st.selectbox("Subject", ["All"] + SUBJECTS)
            df = rev if subj_filter2 == "All" else rev[rev["Subject"]==subj_filter2]
            st.dataframe(df.reset_index(drop=True), use_container_width=True)
